from django.shortcuts import render
from reservation.models import Reservation
from vehicles.models import Vehicle
from client.models import Client
from employee.models import Employee
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from django.shortcuts import render, get_object_or_404
from django.db.models import Count, Sum, Min, Max, Q
from django.utils.timezone import make_aware, datetime
from django.contrib.auth.decorators import login_required
from maintenance.models import Maintenance
# Create your views here.

@login_required
def report_index(request):
    total_vehicles = Vehicle.objects.count()
    available_vehicles = Vehicle.objects.filter(status='available').count()
    
    if total_vehicles > 0:
        available_percent = int((available_vehicles / total_vehicles) * 100)
    else:
        available_percent = 0

    data = {
        'active_clients': Client.objects.filter(active=True).count(),
        'inactive_clients': Client.objects.filter(active=False).count(),
        'total_vehicles': total_vehicles,
        'available_vehicles': available_vehicles,
        'available_percent': available_percent, 
        'total_reservations': Reservation.objects.count(),
        'active_reservations': Reservation.objects.filter(status='active').count(),
        'completed_reservations': Reservation.objects.filter(status='completed').count(),
        'canceled_reservations': Reservation.objects.filter(status='canceled').count(),
        'total_employees': Employee.objects.count(),
        'active_employees': Employee.objects.filter(active=True).count(),
        'inactive_employees': Employee.objects.filter(active=False).count(),
    }
    return render(request, 'reports/index.html', data)

@login_required
def export_PDF(request):
    total_vehicles = Vehicle.objects.count()
    available_vehicles = Vehicle.objects.filter(status='available').count()

    if total_vehicles > 0:
        available_percent = int((available_vehicles / total_vehicles) * 100)
    else:
        available_percent = 0

    data = {
        'Clientes Ativos': Client.objects.filter(active=True).count(),
        'Clientes Inativos': Client.objects.filter(active=False).count(),
        'Total de Veiculos': total_vehicles,
        'Veiculos Disponiveis': available_vehicles,
        'Percentual Disponivel': f'{available_percent}%',
        'Total de Reservas': Reservation.objects.count(),
        'Reservas Ativas': Reservation.objects.filter(status='active').count(),
        'Reservas Completas': Reservation.objects.filter(status='completed').count(),
        'Reservas Canceladas': Reservation.objects.filter(status='canceled').count(),
        'Total de Funcionarios': Employee.objects.count(),
        'Funcionarios Ativos': Employee.objects.filter(active=True).count(),
        'Funcionarios Inativos': Employee.objects.filter(active=False).count(),
    }

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 50, "Relatorio Geral")

    y = height - 100

    pdf.setFont("Helvetica", 12)
    for key, value in data.items():
        pdf.drawString(50, y, f"{key}: {value}")
        y -= 20
        if y < 50:
            pdf.showPage()
            pdf.setFont("Helvetica", 12)
            y = height - 50

    pdf.setFont("Helvetica-Oblique", 10)
    pdf.drawString(50, 30, "Rent a Car")

    pdf.showPage()
    pdf.save()

    return response

@login_required
def export_excel(request):
    total_vehicles = Vehicle.objects.count()
    available_vehicles = Vehicle.objects.filter(status='available').count()
    available_percent = int((available_vehicles / total_vehicles) * 100) if total_vehicles > 0 else 0

    data = {
        'Clientes Ativos': Client.objects.filter(active=True).count(),
        'Clientes Inativos': Client.objects.filter(active=False).count(),
        'Veículos Totais': total_vehicles,
        'Veículos Disponíveis': available_vehicles,
        'Disponibilidade atual de veículos (%)': available_percent,
        'Reservas Totais': Reservation.objects.count(),
        'Reservas Ativas': Reservation.objects.filter(status='active').count(),
        'Reservas Concluídas': Reservation.objects.filter(status='completed').count(),
        'Reservas Canceladas': Reservation.objects.filter(status='canceled').count(),
        'Funcionários Totais': Employee.objects.count(),
        'Funcionários Ativos': Employee.objects.filter(active=True).count(),
        'Funcionários Inativos': Employee.objects.filter(active=False).count(),
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Geral"

    ws.append(["Descrição", "Quantidade"])

    for key, value in data.items():
        ws.append([key, value])

    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="relatorio_dashboard.xlsx"'
    wb.save(response)
    return response

@login_required
def client_historic_report(request):
    clients = Client.objects.all()
    client = None
    reservations = []
    total_reservations = 0
    status_summary = {'active': 0, 'cancelled': 0, 'picked_up': 0, 'returned': 0}
    total_amount_spent = 0
    first_reservation = None
    last_reservation = None

    client_id = request.POST.get('client_id')
    start_date_str = request.POST.get('start_date')
    end_date_str = request.POST.get('end_date')

    if client_id and start_date_str and end_date_str:
        client = get_object_or_404(Client, id=client_id)

        start_date = make_aware(datetime.strptime(start_date_str, "%Y-%m-%d"))
        end_date = make_aware(datetime.strptime(end_date_str, "%Y-%m-%d"))

        reservations = Reservation.objects.filter(
            client=client,
            created_at__range=[start_date, end_date]
        )

        total_reservations = reservations.count()

        status_counts = reservations.values('status').annotate(count=Count('id'))

        for item in status_counts:
            status_summary[item['status']] = item['count']

        for reservation in reservations:
            rented_days = (reservation.end_date - reservation.start_date).days
            daily_price = reservation.vehicle.vehicle_class.daily_price
            total_amount_spent += rented_days * daily_price

        first_reservation = reservations.aggregate(first=Min('start_date'))['first']
        last_reservation = reservations.aggregate(last=Max('start_date'))['last']

    context = {
        'clients': clients,
        'client': client,
        'reservations': reservations,
        'total_reservations': total_reservations,
        'status_summary': status_summary,
        'total_amount_spent': total_amount_spent,
        'first_reservation': first_reservation,
        'last_reservation': last_reservation,
        'client_id': client_id,
        'start_date': start_date_str,
        'end_date': end_date_str,
    }

    return render(request, 'reports/client_historic_report.html', context)

@login_required
def export_client_historic_report_pdf(request):
    client_id = request.GET.get('client_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not client_id or not start_date or not end_date:
        return HttpResponse("Parâmetros inválidos.", status=400)

    client = get_object_or_404(Client, id=client_id)

    reservations = Reservation.objects.filter(
        client=client,
        start_date__range=[start_date, end_date]
    )

    total_reservations = reservations.count()

    status_summary = {'active': 0, 'cancelled': 0, 'picked_up': 0, 'returned': 0}
    status_counts = reservations.values('status').annotate(count=Count('id'))

    for item in status_counts:
        status_summary[item['status']] = item['count']

    total_amount_spent = 0
    for reservation in reservations:
        rented_days = (reservation.end_date - reservation.start_date).days
        daily_price = reservation.vehicle.vehicle_class.daily_price
        total_amount_spent += rented_days * daily_price

    first_reservation = reservations.aggregate(
        first=Min('start_date')
    )['first']

    last_reservation = reservations.aggregate(
        last=Max('start_date')
    )['last']

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_historico_cliente.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    pdf.setTitle("Relatório Histórico de Clientes")
    cm = 28.35  
    y = height - 2 * cm

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(2 * cm, y, "Relatório Histórico de Cliente")
    y -= 1 * cm

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(2 * cm, y, "Dados do Cliente:")
    pdf.setFont("Helvetica", 11)
    y -= 0.7 * cm
    pdf.drawString(2.5 * cm, y, f"Nome: {client.name}")
    y -= 0.5 * cm
    pdf.drawString(2.5 * cm, y, f"CPF: {client.cpf}")
    y -= 0.5 * cm
    pdf.drawString(2.5 * cm, y, f"E-mail: {client.email}")
    y -= 0.5 * cm
    pdf.drawString(2.5 * cm, y, f"Telefone: {client.phone}")
    y -= 0.7 * cm

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(2 * cm, y, "Resumo das Reservas:")
    pdf.setFont("Helvetica", 11)
    y -= 0.7 * cm
    pdf.drawString(2.5 * cm, y, f"Total de Reservas: {total_reservations}")
    y -= 0.5 * cm
    pdf.drawString(2.5 * cm, y, f"Ativas: {status_summary['active']}")
    y -= 0.5 * cm
    pdf.drawString(2.5 * cm, y, f"Retiradas: {status_summary['picked_up']}")
    y -= 0.5 * cm
    pdf.drawString(2.5 * cm, y, f"Retornadas: {status_summary['returned']}")
    y -= 0.5 * cm
    pdf.drawString(2.5 * cm, y, f"Canceladas: {status_summary['cancelled']}")
    y -= 0.5 * cm
    pdf.drawString(2.5 * cm, y, f"Total Gasto: R$ {total_amount_spent:.2f}")
    y -= 0.5 * cm
    pdf.drawString(2.5 * cm, y, f"Primeira Reserva: {first_reservation.strftime('%d/%m/%Y') if first_reservation else '-'}")
    y -= 0.5 * cm
    pdf.drawString(2.5 * cm, y, f"Última Reserva: {last_reservation.strftime('%d/%m/%Y') if last_reservation else '-'}")
    y -= 1 * cm

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(2 * cm, y, "Detalhes das Reservas:")
    y -= 0.7 * cm

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(2 * cm, y, "ID")
    pdf.drawString(3.5 * cm, y, "Veículo")
    pdf.drawString(8.5 * cm, y, "Início")
    pdf.drawString(11 * cm, y, "Fim")
    pdf.drawString(13.5 * cm, y, "Status")
    pdf.drawString(16 * cm, y, "Valor")
    y -= 0.5 * cm
    pdf.line(2 * cm, y + 0.2 * cm, 19 * cm, y + 0.2 * cm)

    pdf.setFont("Helvetica", 9)

    for r in reservations:
        if y <= 3 * cm:
            pdf.showPage()
            y = height - 2 * cm

        rented_days = (r.end_date - r.start_date).days
        daily_price = r.vehicle.vehicle_class.daily_price
        amount = rented_days * daily_price

        pdf.drawString(2 * cm, y, str(r.id))
        pdf.drawString(3.5 * cm, y, str(r.vehicle.plate)[:25])
        pdf.drawString(8.5 * cm, y, r.start_date.strftime('%d/%m/%Y'))
        pdf.drawString(11 * cm, y, r.end_date.strftime('%d/%m/%Y'))
        pdf.drawString(13.5 * cm, y, r.get_status_display())
        pdf.drawString(16 * cm, y, f"R$ {amount:.2f}")
        y -= 0.5 * cm

    pdf.showPage()
    pdf.save()

    return response

@login_required
def export_client_historic_report_excel(request):
    client_id = request.GET.get('client_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not client_id or not start_date or not end_date:
        return HttpResponse("Parâmetros inválidos.", status=400)

    client = get_object_or_404(Client, id=client_id)

    reservations = Reservation.objects.filter(
        client=client,
        start_date__range=[start_date, end_date]
    )

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_historico_cliente_{client.id}.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Histórico de Reservas"

    headers = ["ID", "Veículo", "Início", "Fim", "Status", "Valor"]
    worksheet.append(headers)

    for r in reservations:
        rented_days = (r.end_date - r.start_date).days
        daily_price = r.vehicle.vehicle_class.daily_price
        amount = rented_days * daily_price

        worksheet.append([
            r.id,
            r.vehicle.plate,
            r.start_date.strftime('%d/%m/%Y'),
            r.end_date.strftime('%d/%m/%Y'),
            r.get_status_display(),
            f"R$ {amount:.2f}"
        ])

    workbook.save(response)
    return response

@login_required
def vehicle_historic_report(request):
    vehicles = Vehicle.objects.all()
    vehicle_id = request.POST.get('vehicle_id')
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')

    context = {
        'vehicles': vehicles,
        'vehicle_id': vehicle_id,
        'start_date': start_date,
        'end_date': end_date,
    }

    if vehicle_id and start_date and end_date:
        vehicle = get_object_or_404(Vehicle, id=vehicle_id)
        start = make_aware(datetime.strptime(start_date, "%Y-%m-%d"))
        end = make_aware(datetime.strptime(end_date, "%Y-%m-%d"))

        reservations = Reservation.objects.filter(vehicle=vehicle, created_at__range=(start, end))
        maintenances = Maintenance.objects.filter(
            vehicle=vehicle
        ).filter(
            Q(start_date__range=(start, end)) |
            Q(end_date__range=(start, end))
        )

        status_summary = {
            'active': reservations.filter(status='active').count(),
            'picked_up': reservations.filter(status='picked_up').count(),
            'returned': reservations.filter(status='returned').count(),
            'canceled': reservations.filter(status='canceled').count(),
        }

        total_amount_collected = sum(
            (r.end_date - r.start_date).days * r.vehicle.vehicle_class.daily_price for r in reservations
        )

        context.update({
            'vehicle': vehicle,
            'reservations': reservations,
            'status_summary': status_summary,
            'total_reservations': reservations.count(),
            'total_amount_collected': total_amount_collected,
            'maintenance_count': maintenances.count(),
            'last_maintenance_date': maintenances.order_by('-start_date').first().start_date if maintenances.exists() else None,
            'maintenances': maintenances,
        })

    return render(request, 'reports/vehicle_historic_report.html', context)

@login_required
def export_vehicle_historic_report_pdf(request):
    vehicle_id = request.GET.get('vehicle_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not vehicle_id or not start_date or not end_date:
        return HttpResponse("Parâmetros inválidos.", status=400)

    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    start = make_aware(datetime.strptime(start_date, "%Y-%m-%d"))
    end = make_aware(datetime.strptime(end_date, "%Y-%m-%d"))

    reservations = Reservation.objects.filter(vehicle=vehicle, created_at__range=(start, end))
    maintenances = Maintenance.objects.filter(vehicle=vehicle, start_date__range=(start, end))

    total_rent = sum(
        (r.end_date - r.start_date).days * r.vehicle.vehicle_class.daily_price for r in reservations
    )

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=relatorio_veiculo_{vehicle.plate}.pdf'

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 50

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(50, y, "Relatório Histórico de Veículo")
    y -= 30

    pdf.setFont("Helvetica", 12)
    pdf.drawString(50, y, f"Placa: {vehicle.plate} - {vehicle.brand} {vehicle.model} ({vehicle.year})")
    y -= 20
    pdf.drawString(50, y, f"Período: {start_date} a {end_date}")
    y -= 30

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, y, "Reservas:")
    y -= 20
    pdf.setFont("Helvetica", 10)

    for r in reservations:
        if y < 80:
            pdf.showPage()
            y = height - 50
        valor = (r.end_date - r.start_date).days * r.vehicle.vehicle_class.daily_price
        pdf.drawString(50, y, f"ID {r.id} - {r.start_date.strftime('%d/%m/%Y')} até {r.end_date.strftime('%d/%m/%Y')} | Status: {r.get_status_display()} | Valor: R$ {valor:.2f}")
        y -= 15

    y -= 20
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, y, "Manutenções:")
    y -= 20
    pdf.setFont("Helvetica", 10)

    for m in maintenances:
        if y < 80:
            pdf.showPage()
            y = height - 50
        status = "Concluída" if m.completed else "Em andamento"
        pdf.drawString(50, y, f"{m.start_date.strftime('%d/%m/%Y')} - {m.reason[:60]}... | Status: {status}")
        y -= 15

    y -= 20
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, y, f"Total Gerado em Reservas: R$ {total_rent:.2f}")

    pdf.showPage()
    pdf.save()
    return response

@login_required
def export_vehicle_historic_report_excel(request):
    vehicle_id = request.GET.get('vehicle_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not vehicle_id or not start_date or not end_date:
        return HttpResponse("Parâmetros inválidos.", status=400)

    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    start = make_aware(datetime.strptime(start_date, "%Y-%m-%d"))
    end = make_aware(datetime.strptime(end_date, "%Y-%m-%d"))

    reservations = Reservation.objects.filter(vehicle=vehicle, created_at__range=(start, end))
    maintenances = Maintenance.objects.filter(vehicle=vehicle, start_date__range=(start, end))

    total_rent = sum(
        (r.end_date - r.start_date).days * r.vehicle.vehicle_class.daily_price for r in reservations
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico do Veículo"

    ws.append(["Veículo", f"{vehicle.plate} - {vehicle.brand} {vehicle.model} ({vehicle.year})"])
    ws.append(["Período", f"{start_date} a {end_date}"])
    ws.append([])
    ws.append(["Reservas"])
    ws.append(["ID", "Data Início", "Data Fim", "Status", "Valor"])

    for r in reservations:
        valor = (r.end_date - r.start_date).days * r.vehicle.vehicle_class.daily_price
        ws.append([
            r.id,
            r.start_date.strftime('%d/%m/%Y'),
            r.end_date.strftime('%d/%m/%Y'),
            r.get_status_display(),
            float(valor)
        ])

    ws.append([])
    ws.append(["Manutenções"])
    ws.append(["Data Início", "Previsão Fim", "Status", "Motivo"])

    for m in maintenances:
        status = "Concluída" if m.completed else "Em andamento"
        ws.append([
            m.start_date.strftime('%d/%m/%Y'),
            m.expected_end_date.strftime('%d/%m/%Y'),
            status,
            m.reason[:60]
        ])

    ws.append([])
    ws.append(["Total gerado em reservas", float(total_rent)])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=relatorio_veiculo_{vehicle.plate}.xlsx'
    wb.save(response)
    return response